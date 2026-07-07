#!/usr/bin/env python
import json
import logging
import copy
from pathlib import Path
from typing import Any, Dict, Optional, Tuple
import openpyxl

# Set up logging to write to both a file (json_to_excel.log) and the console
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("json_to_excel.log", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("json_to_excel")

# Expected columns and their mapping to JSON keys
# Mapping definition: Excel Header -> JSON Key
HEADER_TO_JSON_KEY = {
    "Filename": "filename",  # Special case: mapped to stem of the json filename
    "Consumer Number": "consumer_number",
    "Consumer Name": "name",
    "Consumer Address": "address",
    "Pincode": "pincode",
    "Sanction Load": "sanction_load",
    "Sanction Load Unit": "sanction_load_unit",
    "Total Bill Amount": "total_bill_amount",
    "Bill Date": "bill_date",
    "Consumed Units": "unit_consumed",
    "Arrears": "arrears",
    "Rate per Unit": "rate_per_unit",
    "Net Bill Amount": "bill_amount",
    "Is Combined Bill": "is_combined_bill",
    "Number of Months": "combined_months_count",
}


def normalize_header(header_text: Optional[Any]) -> str:
    """
    Normalizes a header name by stripping whitespace, merging multiple spaces,
    and converting to lowercase. This helps match headers dynamically
    even if spacing or capitalization changes in the template.
    """
    if header_text is None:
        return ""
    # Standardize spaces and make lowercase
    return " ".join(str(header_text).strip().split()).lower()


def find_header_row_and_mapping(sheet: openpyxl.worksheet.worksheet.Worksheet, 
                                 target_headers: Dict[str, str]) -> Tuple[Optional[int], Dict[str, int]]:
    """
    Scans the first 20 rows of the sheet to locate the header row.
    Returns a tuple of (header_row_idx, {normalized_header_name: col_index}).
    We consider a row to be the header row if it contains at least 3 matching headers.
    """
    normalized_targets = {normalize_header(h): h for h in target_headers.keys()}
    
    best_row_idx = None
    best_mapping = {}
    max_matches = 0

    # Scan rows 1 to 20 to dynamically identify the header row
    for row_idx in range(1, 21):
        mapping = {}
        matches = 0
        
        # Iterate over all columns in the current row
        for col_idx in range(1, sheet.max_column + 1):
            cell_val = sheet.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                norm_val = normalize_header(cell_val)
                if norm_val in normalized_targets:
                    # Map the target header key to the 1-based column index
                    mapping[normalized_targets[norm_val]] = col_idx
                    matches += 1
        
        # Pick the row that matches the most target headers (minimum 3 matches to be valid)
        if matches > max_matches and matches >= 3:
            max_matches = matches
            best_row_idx = row_idx
            best_mapping = mapping

    return best_row_idx, best_mapping


def copy_cell_style(src_cell, dest_cell):
    """
    Copies cell formatting (font, border, fill, number_format, alignment, protection)
    from src_cell to dest_cell to preserve template formatting in openpyxl.
    """
    if src_cell.has_style:
        dest_cell.font = copy.copy(src_cell.font)
        dest_cell.border = copy.copy(src_cell.border)
        dest_cell.fill = copy.copy(src_cell.fill)
        dest_cell.number_format = src_cell.number_format
        dest_cell.alignment = copy.copy(src_cell.alignment)
        dest_cell.protection = copy.copy(src_cell.protection)


def format_excel_value(val: Any) -> Any:
    """
    Converts values according to specifications:
    - null (None) or empty value -> "null" (written as string "null")
    - true/false -> Native boolean True/False (renders as TRUE/FALSE in Excel)
    - Any other values -> Raw value
    """
    if val is None or val == "":
        return "null"
    if isinstance(val, bool):
        return val  # Python booleans match Excel's native TRUE/FALSE boolean type
    return val


def values_are_equivalent(val1: Any, val2: Any) -> bool:
    """
    Compares two values for equivalence. If one value is numeric (float or int),
    attempts to convert both to float for comparison. Otherwise compares string
    representations after stripping whitespace.
    """
    if val1 == val2:
        return True
    if val1 is None or val2 is None:
        return val1 == val2
    # Check string 'null' case
    if str(val1).strip().lower() == "null" and str(val2).strip().lower() == "null":
        return True
    try:
        if float(val1) == float(val2):
            return True
    except (ValueError, TypeError):
        pass
    return str(val1).strip() == str(val2).strip()


def main():
    # Configure path locations using pathlib
    base_dir = Path(__file__).resolve().parent
    input_folder = base_dir / "outputs"
    template_path = base_dir / "template.xlsx"
    output_path = base_dir / "test.xlsx"

    logger.info("Starting JSON to Excel conversion process...")

    # Validate input sources
    if not input_folder.exists() or not input_folder.is_dir():
        logger.error(f"Input JSON folder '{input_folder}' does not exist or is not a directory.")
        print("Total files processed: 0")
        print("Total successful: 0")
        print("Total failed: 0")
        return

    if not template_path.exists():
        logger.error(f"Excel template file '{template_path}' not found.")
        print("Total files processed: 0")
        print("Total successful: 0")
        print("Total failed: 0")
        return

    # Scan and filter for JSON files in the input folder
    json_files = sorted(list(input_folder.glob("*.json")))
    total_files = len(json_files)
    
    if total_files == 0:
        logger.info(f"No JSON files found in '{input_folder}'.")
        print("Total files processed: 0")
        print("Total successful: 0")
        print("Total failed: 0")
        return

    # Load test.xlsx if it exists to preserve manual edits, otherwise use template.xlsx
    if output_path.exists():
        wb_source_path = output_path
        logger.info(f"Existing output file found. Loading '{output_path.name}' to preserve manual edits.")
    else:
        wb_source_path = template_path
        logger.info(f"No existing output file found. Loading template '{template_path.name}'.")

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(wb_source_path)
        sheet = wb.active
        logger.info(f"Successfully loaded '{wb_source_path.name}'. Active sheet: '{sheet.title}'")
    except Exception as e:
        logger.critical(f"Failed to read/load '{wb_source_path.name}': {e}")
        print("Total files processed: 0")
        print("Total successful: 0")
        print("Total failed: 0")
        return

    # Detect header row and column mapping dynamically
    header_row, col_mapping = find_header_row_and_mapping(sheet, HEADER_TO_JSON_KEY)
    if header_row is None:
        logger.critical("Could not dynamically locate expected headers in the template.")
        print("Total files processed: 0")
        print("Total successful: 0")
        print("Total failed: 0")
        return

    logger.info(f"Detected header row at index: {header_row}")
    for header, col in col_mapping.items():
        logger.info(f"  Header '{header}' -> Column {col}")

    successful_count = 0
    failed_count = 0

    # Determine starting data row (immediately after the header row)
    start_data_row = header_row + 1
    
    # Store references to cells on the first data row for style copying/preservation
    style_source_cells = {}
    for col_idx in col_mapping.values():
        style_source_cells[col_idx] = sheet.cell(row=start_data_row, column=col_idx)

    # Build a map of existing filenames to their row indices to update/skip properly
    existing_filenames = {}
    filename_col_idx = col_mapping.get("Filename")
    
    if filename_col_idx is not None:
        for row_idx in range(start_data_row, sheet.max_row + 1):
            cell_val = sheet.cell(row=row_idx, column=filename_col_idx).value
            if cell_val is not None:
                fn_str = str(cell_val).strip()
                if fn_str:
                    existing_filenames[fn_str] = row_idx

    # Find the next available row to append any new JSON data
    next_available_row = max(existing_filenames.values()) + 1 if existing_filenames else start_data_row

    # Process all JSON files
    for idx, json_path in enumerate(json_files):
        filename_key = json_path.stem
        logger.info(f"Processing file: {json_path.name}")
        
        try:
            # Read and parse JSON content
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            
            # Determine target row and whether it's a new insertion or an update to an existing row
            if filename_key in existing_filenames:
                current_row = existing_filenames[filename_key]
                is_new_row = False
                logger.info(f"  Row for '{filename_key}' already exists at index {current_row}. Checking cells...")
            else:
                current_row = next_available_row
                next_available_row += 1
                is_new_row = True
                logger.info(f"  No row found for '{filename_key}'. Appending new row at index {current_row}.")
            
            # Fill mapped columns
            for header_name, col_idx in col_mapping.items():
                target_cell = sheet.cell(row=current_row, column=col_idx)
                
                # Fetch value from JSON based on mapping rules
                json_key = HEADER_TO_JSON_KEY[header_name]
                if json_key == "filename":
                    raw_value = filename_key
                else:
                    raw_value = data.get(json_key)

                # Format the JSON value
                formatted_value = format_excel_value(raw_value)

                if is_new_row:
                    # For brand new rows, simply write the formatted value
                    target_cell.value = formatted_value
                    # Copy styling from the template cell
                    if current_row > start_data_row:
                        copy_cell_style(style_source_cells[col_idx], target_cell)
                else:
                    # For existing rows, preserve manual edits
                    current_cell_val = target_cell.value
                    
                    if current_cell_val is None or current_cell_val == "" or str(current_cell_val).strip().lower() == "null":
                        # If the Excel cell is empty/null, write the new JSON value
                        target_cell.value = formatted_value
                        logger.debug(f"    Updating empty cell ({current_row}, col {col_idx}) with value '{formatted_value}'")
                    else:
                        # If the Excel cell has an existing value, verify if it's equivalent
                        if not values_are_equivalent(current_cell_val, formatted_value):
                            # The cell value differs from the JSON value, indicating a potential manual edit. Keep it!
                            logger.info(
                                f"    Preserved manual edit in cell ({current_row}, col {col_idx}) for header '{header_name}': "
                                f"'{current_cell_val}' (JSON value would be '{formatted_value}')"
                            )
                        else:
                            # They are equivalent. We can just set it to make sure it's correct
                            target_cell.value = formatted_value
                            
            successful_count += 1
            
        except json.JSONDecodeError as jde:
            logger.error(f"Failed to parse '{json_path.name}' (Invalid JSON structure): {jde}")
            failed_count += 1
        except Exception as e:
            logger.error(f"Failed to process '{json_path.name}': {e}")
            failed_count += 1

    # Save to output_path without overwriting the original template
    try:
        wb.save(output_path)
        logger.info(f"Conversion complete. Output saved successfully to '{output_path.name}'")
    except Exception as e:
        logger.error(f"Failed to save output Excel file to '{output_path}': {e}")

    # Output execution summary counts
    print(f"Total files processed: {total_files}")
    print(f"Total successful: {successful_count}")
    print(f"Total failed: {failed_count}")


if __name__ == "__main__":
    main()
